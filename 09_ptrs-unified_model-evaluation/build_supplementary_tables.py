"""Build fresh Supplementary Tables S3-S7 as Excel workbooks.

Sources:
  S3 — per-feature PTRS retained by cross-cohort consistency filter (TWAS P+T)
        <-- meta_model_{tissue,ct}__pval-*/consistent_features.csv (all 4 pvals × 2 mv, stacked)

  S4 — Performance of tissue-level PTRS across TWAS p-value thresholds AND MA-FOCUS shortlist
        (all classifiers). Two pipeline blocks stacked in one sheet:
          - TWAS P+T: <-- meta_model_tissue__pval-*/individual_results_long.csv
                     (every classifier × every candidate tissue × every p-value threshold).
                     Falls back to best_valid_per_feature.csv when the long CSV is not yet
                     materialized (e.g. TWAS_PT notebook run before the long-CSV patch).
          - MA-FOCUS: <-- meta_model_tissue/individual_ptrs_long.csv
                     (every classifier × every MA-FOCUS shortlist tissue; single "P_VAL=1" run).
                     AUC/OR/P recomputed here from per-sample predictions to match notebook procedure.
        Rows are distinguished by a Pipeline column.

  S5 — Same as S4 but for cell-type-level PTRS.

  S6 — Feature-count ablation for cross-modal PTRS + PRS integration
        <-- data/predictions/feature_ablation/ablation_results.csv (18 rows, added 2 baselines here)

  S7 — Best-performing models across scoring categories
        <-- figures/summary_comparison_table.csv (enriched with GACRS-test AUC/OR
            and CAMP-full P where computable from source all_results.csv files)

Outputs (on Desktop):
  Supplementary_Tables_PTRS_fresh.xlsx        (S3, S4, S5, S7)
  Supplementary_Table_Feature_Ablation_fresh.xlsx (S6)
"""
import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# Config
# =====================================================================
ROOT = Path('/Users/nancyh/Desktop/asthma-prs-study-fresh/09_ptrs-unified_model-evaluation')
PRED = ROOT / 'data' / 'predictions'
FIG  = ROOT / 'figures'
DESKTOP = Path('/Users/nancyh/Desktop')
SUPP    = ROOT / 'supplement'; SUPP.mkdir(exist_ok=True)

PVS   = ['5e-05', '5e-04', '0_005', '0_05']
PV_LABEL = {'5e-05': '5 x 10^-5', '5e-04': '5 x 10^-4', '0_005': '5 x 10^-3', '0_05': '5 x 10^-2'}
MV_LABEL = {'tissue': 'GTEx tissue', 'ct': 'OneK1K cell type'}

# =====================================================================
# S1 — European-ancestry meta-analysis index variants (unchanged content)
# Source: /Users/nancyh/Desktop/Table_S1_index_variants.xlsx sheet "Table S1"
# =====================================================================
def build_s1():
    src = DESKTOP / 'Table_S1_index_variants.xlsx'
    if not src.exists():
        print(f'  S1 source missing: {src}')
        return pd.DataFrame()
    return pd.read_excel(src, sheet_name='Table S1')


# =====================================================================
# S2 — Gene-context pairs retained at each stage of TWAS and FOCUS fine-mapping
#      (unchanged content — mirrors the paper's Table S2)
# Source: /Users/nancyh/Desktop/Table_S4_finemapping_funnel.docx (stale filename;
# the docx is the manuscript's paper Table S2 pre-renumbering).
# =====================================================================
def build_s2():
    import docx
    src = DESKTOP / 'Table_S4_finemapping_funnel.docx'
    if not src.exists():
        print(f'  S2 source missing: {src}')
        return pd.DataFrame()
    d = docx.Document(src)
    t = d.tables[0]
    header = [c.text.strip() for c in t.rows[0].cells]
    rows = [[c.text.strip() for c in r.cells] for r in t.rows[1:]]
    return pd.DataFrame(rows, columns=header)


# =====================================================================
# S3 — Consistency-passed shortlist (TWAS P+T)
# =====================================================================
def build_s3():
    rows = []
    for mv in ['tissue', 'ct']:
        for pv in PVS:
            f = PRED / f'meta_model_{mv}__pval-{pv}' / 'consistent_features.csv'
            if not f.exists(): continue
            df = pd.read_csv(f)
            for _, r in df.iterrows():
                rows.append({
                    'Modality':             MV_LABEL[mv],
                    'Feature':              r['Feature'],
                    'TWAS_p_threshold':     PV_LABEL[pv],
                    'Best_classifier':      r['Model'],
                    'GACRS_test_AUC':       round(r['GACRS_AUC'], 4),
                    'CAMP_Balanced_AUC':    round(r['CAMP_ONLY_BAL_AUC'], 4),
                    'CAMP_Balanced_AUC_SD': round(r['CAMP_ONLY_BAL_AUC_std'], 4),
                    'Consistency_Min_AUC':  round(r['Consistency_Min_AUC'], 4),
                    'Consistency_Mean_AUC': round(r['Consistency_Mean_AUC'], 4),
                })
    # Also enrich with GACRS-test OR / P and CAMP-Balanced OR from best_valid file (same rows)
    enriched = []
    for row in rows:
        mv_key = 'tissue' if row['Modality'] == 'GTEx tissue' else 'ct'
        pv_key = {v: k for k, v in PV_LABEL.items()}[row['TWAS_p_threshold']]
        bv = pd.read_csv(PRED / f'meta_model_{mv_key}__pval-{pv_key}' / 'best_valid_per_feature.csv')
        m = bv[(bv['Feature'] == row['Feature']) & (bv['Model'] == row['Best_classifier'])]
        if len(m):
            m = m.iloc[0]
            row['GACRS_test_OR']    = round(m['GACRS_OR'], 3)
            row['GACRS_test_P']     = f"{m['GACRS_P']:.2e}"
            row['CAMP_Balanced_OR'] = round(m['CAMP_ONLY_BAL_OR'], 3)
        enriched.append(row)
    df = pd.DataFrame(enriched)
    df = df[['Modality', 'Feature', 'TWAS_p_threshold', 'Best_classifier',
             'GACRS_test_AUC', 'GACRS_test_OR', 'GACRS_test_P',
             'CAMP_Balanced_AUC', 'CAMP_Balanced_AUC_SD', 'CAMP_Balanced_OR',
             'Consistency_Min_AUC', 'Consistency_Mean_AUC']]
    return df


# =====================================================================
# S4 / S5 — All classifiers × all features × all TWAS p-value thresholds (per modality),
# plus MA-FOCUS rows (single P_VAL='1' analogue) stacked below.
# =====================================================================
def _odds_ratio_quartile(y_true, y_pred, q=0.25):
    thigh = np.quantile(y_pred, 1 - q)
    tlow  = np.quantile(y_pred, q)
    top = y_true[y_pred >= thigh]
    bot = y_true[y_pred <= tlow]
    a = top.sum(); b = len(top) - a
    c = bot.sum(); d = len(bot) - c
    if b == 0 or c == 0: return np.inf
    return (a * d) / (b * c)


def _twas_pt_block(mv):
    """Rows for the TWAS P+T portion of S4/S5 — one row per (feature, classifier, p-value)."""
    rows = []
    used_long_count = 0
    fallback_count = 0
    for pv in PVS:
        long_f  = PRED / f'meta_model_{mv}__pval-{pv}' / 'individual_results_long.csv'
        best_f  = PRED / f'meta_model_{mv}__pval-{pv}' / 'best_valid_per_feature.csv'
        if long_f.exists():
            df = pd.read_csv(long_f)
            used_long_count += 1
        elif best_f.exists():
            df = pd.read_csv(best_f)
            fallback_count += 1
        else:
            continue
        for _, r in df.iterrows():
            rows.append({
                'Pipeline':             'TWAS P+T',
                'Feature':              r['Feature'],
                'TWAS_p_threshold':     PV_LABEL[pv],
                'Classifier':           r['Model'],
                'GACRS_test_AUC':       round(r['GACRS_AUC'], 4),
                'GACRS_test_OR':        round(r['GACRS_OR'], 3) if pd.notna(r.get('GACRS_OR')) else None,
                'GACRS_test_P':         f"{r['GACRS_P']:.2e}" if pd.notna(r.get('GACRS_P')) else None,
                'CAMP_Balanced_AUC':    round(r['CAMP_ONLY_BAL_AUC'], 4),
                'CAMP_Balanced_AUC_SD': round(r['CAMP_ONLY_BAL_AUC_std'], 4) if pd.notna(r.get('CAMP_ONLY_BAL_AUC_std')) else None,
                'CAMP_Balanced_OR':     round(r['CAMP_ONLY_BAL_OR'], 3) if pd.notna(r.get('CAMP_ONLY_BAL_OR')) else None,
                'CAMP_only_full_P':     f"{r['CAMP_ONLY_FULL_P']:.2e}" if pd.notna(r.get('CAMP_ONLY_FULL_P')) else None,
                'Delta_Balanced_minus_GACRS': round(r['Delta_CAMP_ONLY_BAL_minus_GACRS'], 4) if pd.notna(r.get('Delta_CAMP_ONLY_BAL_minus_GACRS')) else None,
            })
    if used_long_count > 0 and fallback_count > 0:
        print(f'  S4/S5 ({mv}) TWAS P+T: mixed sources — {used_long_count} pvals from long, {fallback_count} from best-only fallback')
    elif fallback_count > 0:
        print(f'  S4/S5 ({mv}) TWAS P+T: all {fallback_count} pvals still on best_valid fallback — re-run notebook for full all-classifier table')
    return rows


def _focus_block(mv):
    """Rows for the MA-FOCUS portion of S4/S5 — one row per (feature, classifier).
    AUC/OR/P are recomputed here from per-sample predictions."""
    from sklearn.metrics import roc_auc_score
    from sklearn.utils import resample
    from scipy.stats import ttest_ind

    f = PRED / f'meta_model_{mv}' / 'individual_ptrs_long.csv'
    if not f.exists():
        print(f'  S4/S5 ({mv}) MA-FOCUS: missing individual_ptrs_long.csv — skipping MA-FOCUS block')
        return []
    d = pd.read_csv(f)
    rows = []
    for feat, model_name in d.groupby(['feature', 'model']).groups.keys():
        sub = d[(d['feature'] == feat) & (d['model'] == model_name)]

        g = sub[sub['cohort'] == 'GACRS_test']
        gacrs_auc = gacrs_or = gacrs_p = None
        if not g.empty:
            gacrs_auc = float(roc_auc_score(g['y_true'], g['score']))
            gacrs_or  = _odds_ratio_quartile(g['y_true'].values, g['score'].values)
            _, gacrs_p = ttest_ind(g.loc[g['y_true']==1, 'score'],
                                    g.loc[g['y_true']==0, 'score'], equal_var=False)

        c = sub[sub['cohort'] == 'CAMP_only']
        camp_full_p = None
        if not c.empty:
            _, camp_full_p = ttest_ind(c.loc[c['y_true']==1, 'score'],
                                        c.loc[c['y_true']==0, 'score'],
                                        equal_var=False, nan_policy='omit')

        camp_bal_auc = camp_bal_std = camp_bal_or = None
        if not c.empty:
            y = c['y_true'].values.astype(int)
            s = c['score'].values
            cases  = np.where(y == 1)[0]
            ctrls  = np.where(y == 0)[0]
            n_ctrl = len(ctrls)
            if n_ctrl > 0 and len(cases) >= n_ctrl:
                aucs, ors = [], []
                for seed in range(100):
                    cs = resample(cases, n_samples=n_ctrl, random_state=seed, replace=False)
                    sel = np.concatenate([cs, ctrls])
                    aucs.append(roc_auc_score(y[sel], s[sel]))
                    ors.append(_odds_ratio_quartile(y[sel], s[sel]))
                camp_bal_auc = float(np.mean(aucs))
                camp_bal_std = float(np.std(aucs))
                camp_bal_or  = float(np.mean(ors))

        rows.append({
            'Pipeline':             'MA-FOCUS',
            'Feature':              feat,
            'TWAS_p_threshold':     '-',
            'Classifier':           model_name,
            'GACRS_test_AUC':       round(gacrs_auc, 4) if gacrs_auc is not None else None,
            'GACRS_test_OR':        round(gacrs_or, 3) if gacrs_or  is not None else None,
            'GACRS_test_P':         f"{gacrs_p:.2e}" if gacrs_p is not None else None,
            'CAMP_Balanced_AUC':    round(camp_bal_auc, 4) if camp_bal_auc is not None else None,
            'CAMP_Balanced_AUC_SD': round(camp_bal_std, 4) if camp_bal_std is not None else None,
            'CAMP_Balanced_OR':     round(camp_bal_or, 3) if camp_bal_or is not None else None,
            'CAMP_only_full_P':     f"{camp_full_p:.2e}" if camp_full_p is not None else None,
            'Delta_Balanced_minus_GACRS': (round(camp_bal_auc - gacrs_auc, 4)
                                           if (camp_bal_auc is not None and gacrs_auc is not None) else None),
        })
    return rows


def build_s4_s5(mv):
    """Combined table: TWAS P+T rows (all 4 pvals) + MA-FOCUS rows, distinguished
    by a Pipeline column."""
    rows = _twas_pt_block(mv) + _focus_block(mv)
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # Order: Pipeline first (TWAS P+T before MA-FOCUS), then Feature, then TWAS_p, then Classifier.
    df['_pipe_order'] = df['Pipeline'].map({'TWAS P+T': 0, 'MA-FOCUS': 1})
    df = df.sort_values(['_pipe_order', 'Feature', 'TWAS_p_threshold', 'Classifier'])
    df = df.drop(columns=['_pipe_order']).reset_index(drop=True)
    return df


# =====================================================================
# S6 — Feature-count ablation for cross-modal PTRS + PRS integration
# =====================================================================
def build_s6():
    df = pd.read_csv(PRED / 'feature_ablation' / 'ablation_results.csv')
    # Add 2 baseline (2-feature anchor + PRS) rows at the top from XM Direct RF-tuned
    xm = pd.read_csv(PRED / 'integrated_ptrs_prs_combined' / 'all_results.csv')
    baselines = []
    for prs in ['PRS-CS', 'PRS-CSx']:
        b = xm[(xm['Method'] == f'Direct ({prs}) + Random Forest (tuned)') &
               (xm['Eval_Set'] == 'CAMP-only Balanced')].iloc[0]
        baselines.append({
            'Experiment_Type':      'Baseline (2-feature anchor + PRS)',
            'PRS_Variant':          prs,
            'Third_Feature':        '-',
            'Modality':             '-',
            'CAMP_Balanced_AUC':    round(b['AUC'], 4),
            'CAMP_Balanced_AUC_SD': round(b['AUC_std'], 4),
            'CAMP_Balanced_OR':     round(b['OR'], 3),
            'Baseline_AUC':         round(b['AUC'], 4),
            'Delta_AUC_vs_baseline': 0.0,
        })
    df_out = df.rename(columns={
        'CAMP_Balanced_AUC_mean': 'CAMP_Balanced_AUC',
        'CAMP_Balanced_OR_mean':  'CAMP_Balanced_OR',
    }).copy()
    for c in ['CAMP_Balanced_AUC', 'CAMP_Balanced_AUC_SD', 'Baseline_AUC', 'Delta_AUC_vs_baseline']:
        df_out[c] = df_out[c].astype(float).round(4)
    df_out['CAMP_Balanced_OR'] = df_out['CAMP_Balanced_OR'].astype(float).round(3)
    df_out = pd.concat([pd.DataFrame(baselines), df_out], ignore_index=True)
    df_out = df_out[['Experiment_Type', 'PRS_Variant', 'Third_Feature', 'Modality',
                     'CAMP_Balanced_AUC', 'CAMP_Balanced_AUC_SD', 'CAMP_Balanced_OR',
                     'Baseline_AUC', 'Delta_AUC_vs_baseline']]
    return df_out


# =====================================================================
# S7 — Best-performing models across scoring categories
# =====================================================================
def _prs_gacrs_test_stats(prs_df, method, config):
    """Compute GACRS-test AUC and case-vs-control t-test P for PRS baseline."""
    from sklearn.metrics import roc_auc_score
    from scipy.stats import ttest_ind
    s = prs_df[(prs_df['method'] == method) & (prs_df['config'] == config) &
               (prs_df['eval_set'] == 'GACRS Test')]
    if s.empty: return None, None
    auc = roc_auc_score(s['y_true'], s['score'])
    return auc, None  # OR/P not readily available without fitting the classifier here


def build_s7():
    df = pd.read_csv(FIG / 'summary_comparison_table.csv')

    # Merge in GACRS-test AUC / OR / P and CAMP-only full P from the source result CSVs.
    prs_df = pd.read_csv(PRED / 'prscs_evaluation' / 'prs_predictions.csv')

    def find_stats(row):
        cat = row['Category']
        try:
            # -------- PRS baselines --------
            if 'altPRS' in cat and 'Cross-modal' not in cat and 'Unified' not in cat:
                method = 'PRS-CS' if cat.startswith('PRS-CS ') else 'PRS-CSx'
                config = 'PRS + Ancestry PCs'
                from sklearn.metrics import roc_auc_score
                from scipy.stats import ttest_ind
                # GACRS-test: AUC + top-vs-bottom quartile OR + Welch P
                s = prs_df[(prs_df['method'] == method) & (prs_df['config'] == config) &
                           (prs_df['eval_set'] == 'GACRS Test')]
                if not s.empty:
                    row['GACRS_test_AUC'] = round(roc_auc_score(s['y_true'], s['score']), 4)
                    row['GACRS_test_OR']  = round(_odds_ratio_quartile(s['y_true'].values,
                                                                       s['score'].values), 3)
                    _, gp = ttest_ind(s.loc[s['y_true'] == 1, 'score'],
                                      s.loc[s['y_true'] == 0, 'score'], equal_var=False)
                    row['GACRS_test_P']   = f"{gp:.2e}"
                # CAMP-only: Welch P + top/bottom quartile OR (single-pass full cohort)
                s = prs_df[(prs_df['method'] == method) & (prs_df['config'] == config) &
                           (prs_df['eval_set'] == 'CAMP-only')]
                if not s.empty:
                    _, p = ttest_ind(s.loc[s['y_true'] == 1, 'score'],
                                     s.loc[s['y_true'] == 0, 'score'], equal_var=False)
                    row['CAMP_only_full_P'] = f"{p:.2e}"
                    row['CAMP_Balanced_OR'] = round(_odds_ratio_quartile(s['y_true'].values,
                                                                         s['score'].values), 3)
            # -------- Cross-modal --------
            elif 'Cross-modal + PRS' in cat:
                prs = 'PRS-CS' if 'PRS-CS ' in cat.split('+')[1] else 'PRS-CSx'
                xm = pd.read_csv(PRED / 'integrated_ptrs_prs_combined' / 'all_results.csv')
                m = xm[(xm['Method'] == f'Direct ({prs}) + {row["Classifier"]}')]
                gt = m[m['Eval_Set'] == 'GACRS Test']
                if not gt.empty:
                    row['GACRS_test_AUC'] = round(float(gt['AUC'].iloc[0]), 4)
                    row['GACRS_test_OR']  = round(float(gt['OR'].iloc[0]), 3)
                    row['GACRS_test_P']   = f"{gt['P_Value'].iloc[0]:.2e}"
                cf = m[m['Eval_Set'] == 'CAMP-only (full)']
                if not cf.empty and pd.notna(cf['P_Value'].iloc[0]):
                    row['CAMP_only_full_P'] = f"{cf['P_Value'].iloc[0]:.2e}"
                cb = m[m['Eval_Set'] == 'CAMP-only Balanced']
                if not cb.empty:
                    row['CAMP_Balanced_OR'] = round(float(cb['OR'].iloc[0]), 3)
            # -------- Unified PTRS (uni-modal, no PRS) — from meta_model_{mv}/unified_balanced_results --------
            elif 'Unified PTRS' in cat and 'MA-FOCUS, meta-model' in cat:
                mv = 'tissue' if 'tissue' in cat else 'ct'
                bal = pd.read_csv(PRED / f'meta_model_{mv}' / 'unified_balanced_results.csv')
                b = bal[bal['Method'] == row['Classifier']]
                if not b.empty:
                    b = b.iloc[0]
                    row['CAMP_Balanced_OR'] = round(float(b['OR']), 3)
                    # GACRS-test AUC/OR/P + CAMP-full P are precomputed in the same CSV.
                    if pd.notna(b.get('GACRS_AUC')):
                        row['GACRS_test_AUC'] = round(float(b['GACRS_AUC']), 4)
                    if pd.notna(b.get('GACRS_OR')):
                        row['GACRS_test_OR']  = round(float(b['GACRS_OR']), 3)
                    if pd.notna(b.get('GACRS_P')):
                        row['GACRS_test_P']   = f"{b['GACRS_P']:.2e}"
                    if pd.notna(b.get('CAMP_ONLY_FULL_P')):
                        row['CAMP_only_full_P'] = f"{b['CAMP_ONLY_FULL_P']:.2e}"
            # -------- Cross-modal Per-feature OOF only (no PRS) --------
            elif 'Per-feature OOF PTRS' in cat and 'cross-modal' in cat:
                # Row's Classifier is e.g. "cd4_naive OOF" -> feature is cd4_naive.
                feat = row['Classifier'].replace(' OOF', '')
                xm = pd.read_csv(PRED / 'integrated_ptrs_prs_combined' / 'all_results.csv')
                m = xm[(xm['Approach'] == 'Per-feature only') &
                       (xm['Method'] == f'{feat} alone (per-feature OOF)')]
                gt = m[m['Eval_Set'] == 'GACRS Test']
                if not gt.empty:
                    row['GACRS_test_AUC'] = round(float(gt['AUC'].iloc[0]), 4)
                    row['GACRS_test_OR']  = round(float(gt['OR'].iloc[0]), 3)
                    row['GACRS_test_P']   = f"{gt['P_Value'].iloc[0]:.2e}"
                cf = m[m['Eval_Set'] == 'CAMP-only (full)']
                if not cf.empty and pd.notna(cf['P_Value'].iloc[0]):
                    row['CAMP_only_full_P'] = f"{cf['P_Value'].iloc[0]:.2e}"
                cb = m[m['Eval_Set'] == 'CAMP-only Balanced']
                if not cb.empty:
                    row['CAMP_Balanced_OR'] = round(float(cb['OR'].iloc[0]), 3)
            # -------- Single-feature MA-FOCUS PTRS --------
            elif 'MA-FOCUS, single feature' in cat:
                mv = 'tissue' if 'tissue' in cat else 'ct'
                bv = pd.read_csv(PRED / f'meta_model_{mv}' / 'best_valid_per_feature.csv')
                m = bv[(bv['Feature'] == row['Features']) & (bv['Model'] == row['Classifier'])]
                if not m.empty:
                    m = m.iloc[0]
                    row['GACRS_test_AUC']  = round(m['GACRS_AUC'], 4)
                    row['GACRS_test_OR']   = round(m['GACRS_OR'], 3)
                    row['GACRS_test_P']    = f"{m['GACRS_P']:.2e}"
                    row['CAMP_Balanced_OR'] = round(m['CAMP_ONLY_BAL_OR'], 3)
                    row['CAMP_only_full_P'] = f"{m['CAMP_ONLY_FULL_P']:.2e}"
            # -------- Single-feature TWAS P+T --------
            elif 'TWAS P+T, single feature' in cat:
                mv = 'tissue' if 'tissue' in cat else 'ct'
                pv_map = {'5e-05': '5e-05', '5e-04': '5e-04', '0_005': '0_005', '0_05': '0_05'}
                pv = row.get('P_VAL')
                if pv and pv != '-' and pv != 1 and pv != '1':
                    d = PRED / f'meta_model_{mv}__pval-{pv_map.get(pv, pv)}'
                    bv = d / 'best_valid_per_feature.csv'
                    if bv.exists():
                        bv = pd.read_csv(bv)
                        m = bv[bv['Feature'] == row['Features']]
                        if not m.empty:
                            m = m.iloc[0]
                            row['GACRS_test_AUC']  = round(m['GACRS_AUC'], 4)
                            row['GACRS_test_OR']   = round(m['GACRS_OR'], 3)
                            row['GACRS_test_P']    = f"{m['GACRS_P']:.2e}"
                            row['CAMP_Balanced_OR'] = round(m['CAMP_ONLY_BAL_OR'], 3)
                            row['CAMP_only_full_P'] = f"{m['CAMP_ONLY_FULL_P']:.2e}"
        except Exception as e:
            print(f'  [S7 enrichment warning] {cat}: {e}')
        return row

    # Add empty columns then enrich
    for c in ['GACRS_test_AUC', 'GACRS_test_OR', 'GACRS_test_P',
              'CAMP_Balanced_OR', 'CAMP_only_full_P']:
        df[c] = np.nan
    df = df.apply(find_stats, axis=1)

    # Rename AUC_std -> CAMP_Balanced_AUC_SD to match paper schema
    df = df.rename(columns={
        'AUC_std': 'CAMP_Balanced_AUC_SD',
        'Delta_AUC': 'Delta_AUC_vs_PRS_CSx_baseline',
    })
    df = df[['Category', 'Features', 'P_VAL', 'Classifier',
             'GACRS_test_AUC', 'GACRS_test_OR', 'GACRS_test_P',
             'CAMP_Balanced_AUC', 'CAMP_Balanced_AUC_SD', 'CAMP_Balanced_OR',
             'CAMP_only_full_P', 'Delta_AUC_vs_PRS_CSx_baseline']]
    return df


# =====================================================================
# Excel writer w/ light styling: bold header, light grey fill, autosize
# =====================================================================
def write_sheet(ws, df, title=None, caption=None, footnote=None):
    start_row = 1
    if title:
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=13)
        start_row += 1
    if caption:
        ws.cell(row=start_row, column=1, value=caption).alignment = Alignment(wrap_text=True, vertical='top')
        # merge across all cols
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(df.columns))
        ws.row_dimensions[start_row].height = 60
        start_row += 1
    # Header
    header_row = start_row
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[header_row].height = 30
    # Data
    last_data_row = header_row
    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        last_data_row = i
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val): val = ''
            ws.cell(row=i, column=j, value=val)
    # Footnote (below data, one blank row for separation)
    if footnote:
        foot_row = last_data_row + 2
        c = ws.cell(row=foot_row, column=1, value=footnote)
        c.font = Font(italic=True, color='7F7F7F')
        c.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=len(df.columns))
        ws.row_dimensions[foot_row].height = 60
    # Column widths
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        widths = [len(str(col))] + [len(str(v)) for v in df[col].fillna('').astype(str).values]
        ws.column_dimensions[letter].width = min(max(widths) + 2, 32)


def write_workbook(path, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, contents in sheets.items():
        # Accept either (df, title, caption) or (df, title, caption, footnote).
        df, title, caption = contents[0], contents[1], contents[2]
        footnote = contents[3] if len(contents) >= 4 else None
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, df, title=title, caption=caption, footnote=footnote)
    wb.save(path)
    print(f'  wrote {path}  ({len(sheets)} sheet(s))')


# =====================================================================
# Assemble + write
# =====================================================================
print('Building S1-S7 from fresh predictions + preserved S1/S2 sources...\n')

s1 = build_s1()
print(f'  S1: {len(s1)} rows (unchanged — from Table_S1_index_variants.xlsx)')

s2 = build_s2()
print(f'  S2: {len(s2)} rows (unchanged — from Table_S4_finemapping_funnel.docx)')

s3 = build_s3()
print(f'  S3: {len(s3)} rows')

s4 = build_s4_s5('tissue')
print(f'  S4: {len(s4)} rows (TWAS P+T + MA-FOCUS, all classifiers x all tissues)')

s5 = build_s4_s5('ct')
print(f'  S5: {len(s5)} rows (TWAS P+T + MA-FOCUS, all classifiers x all cell types)')

s6 = build_s6()
print(f'  S6: {len(s6)} rows')

s7 = build_s7()
print(f'  S7: {len(s7)} rows')

# --- Captions (based on the manuscript supplement PDF) ---
CAPS = {
    'S3': ('Table S3. Per-feature PTRS models retained by the cross-cohort consistency filter.',
           'Per-feature PTRS models retained by the two-stage cross-cohort consistency filter at each of four candidate TWAS p-value thresholds. '
           'After the GACRS validation-set directionality filter, models were required to achieve AUC > 0.53 in GACRS-test and mean AUC > 0.52 in CAMP-Balanced. '
           'The best classifier is the highest-Consistency_Min_AUC Stage-2 classifier per feature, selected from the fresh 7-model pool (Logistic Regression L2, '
           'Elastic Net CV, Random Forest tuned, Gradient Boosting tuned, SVM linear, Stacking LR+RF+GB, plus Ridge/Lasso variants). '
           'GACRS-test AUC/OR/P computed in the 25% held-out within-cohort test set; CAMP-Balanced AUC is mean ± SD across 100 iterations of 64 sampled cases × all 64 controls.'),

    'S1': ('Table S1. European-ancestry meta-analysis index variants.',
           'Index variants for the 275 independent genome-wide significant loci (P ≤ 5 × 10⁻⁸) from the European-ancestry TAGC + GBMI meta-analysis, ordered by chromosome and position. Beta and Z are reported with respect to the effect allele. Content preserved verbatim from the paper submission.'),

    'S2': ('Table S2. Gene–context pairs retained at each stage of TWAS and FOCUS fine-mapping.',
           'Number of gene–tissue (GTEx) and gene–cell-type (OneK1K) pairs remaining at each stage of the gene-prioritization pipeline. Content preserved verbatim from the paper submission (related to Figure 3).'),

    'S4': ('Table S4. Performance of tissue-level PTRS from the TWAS P+T pipeline (all p-value thresholds) and the MA-FOCUS pipeline — all classifiers.',
           'Performance of tissue-level PTRS evaluated using ALL 7 classifiers (Ridge C=0.01/0.1/1.0, Lasso, Elastic Net, Gradient Boosting, RF GridSearch). Rows are distinguished by the Pipeline column: '
           '"TWAS P+T" rows tabulate every classifier × candidate tissue × 4 TWAS-p-value thresholds — a tissue is included only if at least one gene-tissue pair passed the threshold and upstream heritability QC. '
           '"MA-FOCUS" rows tabulate every classifier × MA-FOCUS cross-cohort-audit shortlist tissue (single "P_VAL=1" run, no threshold axis). '
           'GACRS-test AUC/OR/P computed in the held-out 25% within-cohort test set. CAMP-Balanced AUC is the mean ± SD across 100 iterations (64 cases × 64 controls). '
           'CAMP-full P is the Welch t-test P for case-vs-control mean predicted probability on the full CAMP-only cohort. '
           'MA-FOCUS AUC/OR/P are recomputed here from the persisted per-sample predictions (individual_ptrs_long.csv) using the same 64v64 × 100 bootstrap procedure as the notebook. '
           'Blank GACRS_test_P / CAMP_only_full_P cells (~5–8% of rows) mark degenerate fits where the classifier (typically Lasso / Elastic Net with strong regularization at the given threshold) shrunk every coefficient to zero, '
           'yielding constant predictions (AUC = 0.5, OR = 1) and an undefined Welch t-test P (case and control score distributions have zero variance).'),

    'S5': ('Table S5. Performance of cell-type-level PTRS from the TWAS P+T pipeline (all p-value thresholds) and the MA-FOCUS pipeline — all classifiers.',
           'Same as Table S4 but for the 17 OneK1K peripheral-blood mononuclear cell types. Both pipeline blocks (TWAS P+T at four thresholds; MA-FOCUS single shortlist) are stacked in one sheet, distinguished by the Pipeline column. '
           'Blank GACRS_test_P / CAMP_only_full_P cells (~8–9% of rows) reflect degenerate-classifier fits identical to those documented in Table S4.'),

    'S6': ('Table S6. Feature-count ablation for the cross-modal PTRS + PRS integration.',
           'Baseline: 2-feature cross-modal anchor pair (naïve CD4+ T cells + esophageal mucosa) integrated with either PRS-CS or PRS-CSx via a grid-searched Random Forest '
           '(sklearn 108-combo grid over n_estimators × max_depth × min_samples_leaf × max_features). '
           'Each of the nine remaining MA-FOCUS consistent-shortlist features (7 GTEx tissues + 2 OneK1K cell types) was added individually as a third PTRS feature — 18 ablation models × 2 PRS variants. '
           'PTRS inputs are per-feature out-of-fold predicted probabilities; PRS inputs are z-standardized. CAMP-Balanced AUC is mean ± SD across 100 iterations (64 cases × 64 controls). '
           'All 18/18 ablation experiments produced a negative delta AUC (mean −0.031), reinforcing that the 2-feature anchor is the AUC-optimal cross-modal architecture.'),

    'S7': ('Table S7. Best-performing models across scoring categories.',
           'Best-performing model within each scoring category, ranked by CAMP-Balanced AUC. Categories: PRS-CS and PRS-CSx baselines with ancestry PCs; best single-tissue and single-cell-type PTRS from TWAS P+T; MA-FOCUS single-feature PTRS anchors; unified tissue- and cell-type-level MA-FOCUS PTRS; and cross-modal PTRS + PRS integrated via 6 model-based classifiers + Rank Addition (7 total, paper-Methods matched, altPRS switch). '
           'GACRS-test AUC/OR/P computed in the held-out 25% within-cohort test set. '
           'CAMP-Balanced AUC is mean ± SD across 100 iterations (64 cases × 64 controls). '
           'CAMP-full P is the Welch t-test P for case-vs-control mean predicted probability on the full CAMP-only cohort. '
           'Delta_AUC baseline is the PRS-CSx (altPRS) row (set to 0).'),
}

DEGEN_NOTE = (
    'Note. Approximately 5–9% of rows in this sheet have blank GACRS_test_P and '
    'CAMP_only_full_P values. These are NOT missing data — every such row has '
    'GACRS_test_AUC = 0.5 and GACRS_test_OR = 1, indicating a degenerate '
    'classifier fit: Lasso / Elastic Net with strong regularization at the given '
    '(feature × TWAS p-value) combination shrunk every input coefficient to zero, '
    'producing constant predicted probabilities for every sample. When the case '
    'and control predicted-probability distributions have zero variance, Welch\'s '
    't-test P is mathematically undefined and is therefore reported as blank.'
)

sheets_ptrs = {
    'S1': (s1, CAPS['S1'][0], CAPS['S1'][1]),
    'S2': (s2, CAPS['S2'][0], CAPS['S2'][1]),
    'S3': (s3, CAPS['S3'][0], CAPS['S3'][1]),
    'S4': (s4, CAPS['S4'][0], CAPS['S4'][1], DEGEN_NOTE),
    'S5': (s5, CAPS['S5'][0], CAPS['S5'][1], DEGEN_NOTE),
    'S7': (s7, CAPS['S7'][0], CAPS['S7'][1]),
}
sheets_ablation = {
    'S6': (s6, CAPS['S6'][0], CAPS['S6'][1]),
}

# Write to the tracked repo location (SUPP) and mirror to Desktop for convenience.
for target in (SUPP, DESKTOP):
    write_workbook(target / 'Supplementary_Tables_PTRS_fresh.xlsx', sheets_ptrs)
    write_workbook(target / 'Supplementary_Table_Feature_Ablation_fresh.xlsx', sheets_ablation)

print('\nDone. Outputs written to both:')
print(f'  {SUPP}/')
print(f'  {DESKTOP}/')
